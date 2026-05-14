import csv
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Count
from surveys.models import Sondage, Question, Choix
from accounts.models import Anonyme, Participant
from .models import Reponse, ReponseDetail

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ─── Helpers ──────────────────────────────────────────────────────────────────

def get_ou_creer_anonyme(request):
    """
    Récupère ou crée un Anonyme via le token de session.
    L'anonyme hérite de Participant (héritage multi-table).
    """
    token = request.session.get('anonyme_token')
    if token:
        try:
            return Anonyme.objects.get(token=token)
        except Anonyme.DoesNotExist:
            pass
    anonyme = Anonyme.objects.create(
        adresse_ip=request.META.get('REMOTE_ADDR', '0.0.0.0'),
    )
    request.session['anonyme_token'] = str(anonyme.token)
    return anonyme


def get_participant(request):
    """
    Retourne le Participant (Utilisateur ou Anonyme) correspondant à la requête.
    Toujours un objet Participant — jamais NULL.
    """
    if request.user.is_authenticated:
        return request.user  # Utilisateur hérite de Participant
    return get_ou_creer_anonyme(request)
    

def a_deja_repondu(sondage, participant):
    """Vérifie si ce participant a déjà soumis une réponse complète à ce sondage"""
    return Reponse.objects.filter(
        sondage=sondage,
        participant=participant,
        est_complete=True
    ).exists()


# ─── Répondre à un sondage ────────────────────────────────────────────────────

def repondre_sondage(request, slug):
    sondage = get_object_or_404(Sondage, slug=slug)

    if not sondage.est_ouvert():
        return render(request, 'surveys/sondage_ferme.html', {'sondage': sondage})

    participant = get_participant(request)

    if sondage.une_seule_reponse and a_deja_repondu(sondage, participant):
        return render(request, 'responses/deja_repondu.html', {'sondage': sondage})

    questions = sondage.question_set.prefetch_related('choix_set').order_by('ordre')

    if request.method == 'POST':
        # Créer la réponse principale
        reponse = Reponse.objects.create(
            sondage=sondage,
            participant=participant,
            est_complete=False,
        )

        # Enregistrer les détails question par question
        for question in questions:
            detail = ReponseDetail.objects.create(
                reponse=reponse,
                question=question,
            )

            if question.type_question in [Question.TYPE_CHOIX_UNIQUE, Question.TYPE_CHOIX_MULTIPLE]:
                choix_ids = request.POST.getlist(f'question_{question.id}')
                if choix_ids:
                    choix_valides = Choix.objects.filter(id__in=choix_ids, question=question)
                    detail.choix_selectionnes.set(choix_valides)

            elif question.type_question in [Question.TYPE_ECHELLE, Question.TYPE_TEXTE]:
                valeur = request.POST.get(f'question_{question.id}', '').strip()
                detail.valeur_texte = valeur
                detail.save()

        # Marquer comme complète
        reponse.est_complete = reponse.est_complete_check()
        reponse.save()

        messages.success(request, "Merci pour votre participation !")
        return redirect('remerciement', slug=sondage.slug)

    return render(request, 'responses/repondre.html', {
        'sondage': sondage,
        'questions': questions,
    })


def remerciement(request, slug):
    sondage = get_object_or_404(Sondage, slug=slug)
    return render(request, 'responses/remerciement.html', {'sondage': sondage})


# ─── Résultats & Statistiques ─────────────────────────────────────────────────

def resultats_sondage(request, slug):
    sondage = get_object_or_404(Sondage, slug=slug)

    # Seul le créateur peut voir les résultats détaillés
    if request.user != sondage.createur and not request.user.is_staff:
        messages.error(request, "Accès refusé aux résultats.")
        return redirect('sondage_detail', slug=slug)

    questions = sondage.question_set.prefetch_related('choix_set').order_by('ordre')
    total_reponses = sondage.reponse_set.filter(est_complete=True).count()
    stats_questions = []

    for question in questions:
        stats = {
            'question': question,
            'type': question.type_question,
            'total': 0,
            'donnees': [],
        }

        if question.type_question in [Question.TYPE_CHOIX_UNIQUE, Question.TYPE_CHOIX_MULTIPLE]:
            choix_stats = []
            for choix in question.obtenir_choix():
                count = choix.reponsedetail_set.count()
                choix_stats.append({
                    'label': choix.texte,
                    'count': count,
                })
            total_q = sum(c['count'] for c in choix_stats)
            for c in choix_stats:
                c['pourcentage'] = round((c['count'] / total_q * 100) if total_q else 0, 1)
            stats['donnees'] = choix_stats
            stats['total'] = total_q

        elif question.type_question == Question.TYPE_ECHELLE:
            details = ReponseDetail.objects.filter(
                question=question,
                reponse__est_complete=True
            ).exclude(valeur_texte='')
            valeurs = [int(d.valeur_texte) for d in details if d.valeur_texte.isdigit()]
            stats['moyenne'] = round(sum(valeurs) / len(valeurs), 2) if valeurs else 0
            stats['distribution'] = [valeurs.count(i) for i in range(1, 6)]
            stats['total'] = len(valeurs)

        elif question.type_question == Question.TYPE_TEXTE:
            details = ReponseDetail.objects.filter(
                question=question,
                reponse__est_complete=True
            ).exclude(valeur_texte='')
            stats['reponses_texte'] = [d.valeur_texte for d in details]
            stats['total'] = details.count()

        stats_questions.append(stats)

    return render(request, 'responses/resultats.html', {
        'sondage': sondage,
        'stats_questions': stats_questions,
        'total_reponses': total_reponses,
    })


def api_stats_question(request, question_id):
    """Endpoint JSON pour Chart.js — données en temps réel"""
    question = get_object_or_404(Question, id=question_id)

    if request.user != question.sondage.createur:
        return JsonResponse({'error': 'Accès refusé'}, status=403)

    data = {'labels': [], 'values': [], 'type': question.type_question}

    if question.type_question in [Question.TYPE_CHOIX_UNIQUE, Question.TYPE_CHOIX_MULTIPLE]:
        for choix in question.obtenir_choix():
            data['labels'].append(choix.texte)
            data['values'].append(choix.compter_reponses())

    elif question.type_question == Question.TYPE_ECHELLE:
        data['labels'] = ['1', '2', '3', '4', '5']
        details = ReponseDetail.objects.filter(question=question, reponse__est_complete=True)
        for i in range(1, 6):
            data['values'].append(details.filter(valeur_texte=str(i)).count())

    return JsonResponse(data)


# ─── Export CSV ───────────────────────────────────────────────────────────────

def export_csv(request, slug):
    sondage = get_object_or_404(Sondage, slug=slug, createur=request.user)
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="sondage_{sondage.slug}.csv"'
    response.write('\ufeff')  # BOM pour Excel

    writer = csv.writer(response)
    questions = list(sondage.question_set.order_by('ordre'))
    entetes = ['Date de réponse', 'Participant'] + [q.texte for q in questions]
    writer.writerow(entetes)

    reponses = sondage.reponse_set.filter(est_complete=True).prefetch_related(
        'reponsedetail_set__choix_selectionnes', 'reponsedetail_set__question'
    )

    for reponse in reponses:
        ligne = [
            reponse.date_envoi.strftime('%d/%m/%Y %H:%M'),
            reponse.get_participant_display(),
        ]
        details_map = {d.question_id: d for d in reponse.reponsedetail_set.all()}
        for question in questions:
            detail = details_map.get(question.id)
            ligne.append(detail.obtenir_affichage() if detail else '—')
        writer.writerow(ligne)

    return response


# ─── Export Excel ─────────────────────────────────────────────────────────────

def export_excel(request, slug):
    if not OPENPYXL_AVAILABLE:
        messages.error(request, "openpyxl non installé. Utilisez l'export CSV.")
        return redirect('resultats_sondage', slug=slug)

    sondage = get_object_or_404(Sondage, slug=slug, createur=request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résultats"

    questions = list(sondage.question_set.order_by('ordre'))
    entetes = ['Date', 'Participant'] + [q.texte for q in questions]
    ws.append(entetes)

    # Style des en-têtes
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')

    reponses = sondage.reponse_set.filter(est_complete=True)
    for reponse in reponses:
        ligne = [
            reponse.date_envoi.strftime('%d/%m/%Y %H:%M'),
            reponse.get_participant_display(),
        ]
        details_map = {d.question_id: d for d in reponse.reponsedetail_set.all()}
        for question in questions:
            detail = details_map.get(question.id)
            ligne.append(detail.obtenir_affichage() if detail else '—')
        ws.append(ligne)

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sondage_{sondage.slug}.xlsx"'
    wb.save(response)
    return response
